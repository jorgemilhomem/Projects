#!/usr/bin/env python
# coding: utf-8

# In[43]:


import os
import tkinter as tk
from tkinter import filedialog
import openpyxl
from distutils.dir_util import copy_tree
import pprint
import shutil

app = tk.Tk()
app.resizable(height = None, width = None)



def action():
    global path_ficha
    path_ficha = filedialog.askopenfilename(title ="Escolher Ficha de Projeto")
    print("path_ficha: ",path_ficha)
    if "FICHA DE PROJETO" in path_ficha:
        # Grava o path da obra numa string
        if "6-DADOS DO PROJETO" in path_ficha:
            path_obra = path_ficha[:len(path_ficha)-29-(len("6-DADOS DO PROJETO")+1)]
        else:
            path_obra = path_ficha[:len(path_ficha)-29]
        print("path_obra: ",path_obra)

        # ABRE A FICHA DE PROJETO.XLSX
        fichaprojeto = openpyxl.load_workbook(path_ficha)
        # NOME (STR) DA PRIMEIRA E UNICA FOLHA DO FICHEIRO
        folha_nome = fichaprojeto.sheetnames[0]
        # PRIMEIRA FOLHA
        folha = fichaprojeto[folha_nome]
        fase = []
        especialidade = []
        dicionario = {}
        #PERCORRE AS CÉLULAS DE INTERESSE NA FOLHA
        for row in range(1, folha.max_row + 1):
            for column in "ABCDEFGHIJKLMNOPQRSTUV":
                # CONSTRÓI STRING COM O NOME DA CÉLULA P SER USADA COMO ÍNDICE
                cell_name = "{}{}".format(column, row)
                # ENCONTRA AS CRUZES E CONSTRÓI LISTAS COM AS FASES E ESPECIALIDADES CORRESPONDENTES ÀS CRUZES
                if folha[cell_name].value != None and isinstance(folha[cell_name].value, str) == True and folha[cell_name].value.upper() == 'X':
                    if column in "FGHIJK" and row < 47:
                        fase.append(str(folha[column + "20"].value))
                        especialidade.append(str(folha["A" + str(row)].value).replace("''",""))
                    if column in "QRSTUV" and row < 47:
                        fase.append(str(folha[column + "20"].value))
                        especialidade.append(str(folha["L" + str(row)].value).replace("''",""))
                    if row in range(47,50) and column == "G":
                        fase.append(str(folha["A" + str(row)].value)[5:7])
                        especialidade.append("")
                        print("A" + str(row))
                        print("fase: ", fase)
        print("especialidade: ", especialidade)
        #Cria dicionário fase:especialidades, sem repetição de keys (fase) a partir das listas criadas acima
        lista = list(zip(fase,especialidade))
        from collections import defaultdict
        res = defaultdict(list)
        for i, j in zip(fase,especialidade):
            res[i].append(j)
        dicionario = dict(res)
        pprint.pprint(dicionario)
        
        # Dicionários com os títulos das pastas a serem criadas/copiadas
        dicionario_fases = {'0':'0-PROGRAMA BASE','1':'1-ESTUDO PREVIO','2':'2-ANTEPROJETO','3':'3-LICENCIAMENTO',                           '4':'4-EXECUCAO','5':'5-TELAS FINAIS','09':'9-ASSISTENCIA TECNICA','10':'10-CERTIFICACAO ENERGETICA',                           '11':'11-BIM-WIP'}
        dicionario_especialidades = {'01': '01-ASPIRACAO CENTRAL', '02': '02-ACUSTICO', '03': '03-SCE-REH', '04': '04-SCE-RECS',                                     '05': '05-RSU', '06': '06-FUNDACOES E ESTRUTURA', '07': '07-EQUIPAMENTOS COZINHAS',                                     '08': '08-PSS', '09': '09-ARQUITETURA', '10': '10-AVAC', '11': '11-PPGRCD',                                     '13': '13-REDE AQUECIMENTO CENTRAL', '14': '14-TRABALHOS COMPLEMENTARES',                                     '15': '15-VENTILACAO E EXAUSTAO FUMOS', '16': '16-ENERGIAS ALTERNATIVAS',                                     '20': '20-REDE GAS', '21': '21-INFRAESTRUTURAS EXTERIORES GAS', '22': '22-REDE VAPOR',                                     '23': '23-REDE VACUO', '24': '24-REDE AR COMPRIMIDO', '25': '25-REDE GASOLEO',                                     '26': '26-REDE GASES MEDICINAIS', '30': '30-REDE AGUAS', '31': '31-TRATAMENTO AGUAS PISCINAS',                                     '32': '32-INFRAESTRUTURAS EXTERIORES AGUAS', '33': '33-RIA', '34': '34-REDE DE REGA',                                     '40': '40-REDE ESGOTOS', '41': '41-INFRAESTRUTURAS EXTERIORES ESGOTOS',                                     '42': '42-REDE DE ÁGUAS PLUVIAIS', '50': '50-TRANSPORTE DE PESSOAS E CARGA',                                     '60': '60-INSTALACOES ELETRICAS', '61': '61-POSTO TRANSFORMACAO',                                     '62': '62-GRUPO GERADOR', '63': '63-INFRAESTRUTURAS EXTERIORES ELETRICAS',                                     '65': '65-GESTAO TECNICA CENTRALIZADA', '70': '70-ITED', '71': '71-ITUR',                                     '80': '80-SEGURANCA CONTRA INCENDIOS', '81': '81-INFRAESTRUTURAS SEGURANCA',                                     '82': '82-SEGURANCA INTEGRADA', '83': '83-MAP', '90': '90-SERVICOS AFETADOS',                                     '91': '91-REDES ESPECIAIS', '92': '92-OVP'}


        # Iterações sobre a fase e correspondentes especialidades p criar as pastas
        txtA = tk.Text(app, width=55, height=24, wrap='none')
        for f in sorted(dicionario):
            print("!!!!!!: ",f)
            print(dicionario_fases[f])
            # Cria pastas das fases, ARQUIT e ORIGINAIS
            os.mkdir(path_obra + dicionario_fases[f])
            if f not in ['09','10','11']:
                os.mkdir(path_obra + dicionario_fases[f] + "\\" + "ARQUIT")
                os.mkdir(path_obra + dicionario_fases[f] + "\\" + "ORIGINAIS")

            txtA.insert('end',"_____________________________________________"+ '\n')
            txtA.insert('end',"Pastas copiadas para " + dicionario_fases[f] + ":" + '\n')
            for e in sorted(dicionario[f]):
                if e in dicionario_especialidades:
                    #print(dicionario[f])
                    path_copiar = "\\\\fileserver\\Obras\\Projetos\\Ano2020\\00000_(nao_mexer)\\18000_(Não mexer)\\" + dicionario_fases[f] + '\\' + dicionario_especialidades[e]
                    path_colar = path_obra + '\\' + dicionario_fases[f] + '\\' + dicionario_especialidades[e]

                    shutil.copytree(path_copiar,path_colar)

                    txtA.insert('end',dicionario_especialidades[e] + '\n')
            txtA.pack()


escolheficha_button = tk.Button(app, pady=10, text="Escolher Ficha de Projeto", command = action)
escolheficha_button.pack(pady=20,padx=20)

app.mainloop()