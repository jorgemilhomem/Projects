import tkinter as tk
from tkinter import filedialog
import os
import openpyxl
import pprint
import collections

app = tk.Tk()

app.title('ObrasFinder')
app.geometry('900x450')

os.chdir(r'\\fileserver\Obras\Util\LISTA DE OBRAS')
theFile = openpyxl.load_workbook('mapa-de-obras.xlsx')
allSheetNames = theFile.sheetnames



def find_specific_cell(argv=None):
    num_resultados = 0
    results_dic = {}
    if keyword_text.get() == '':
            return 'Do nothing'
    for sheet in allSheetNames:
       # print("Current sheet name is {}" .format(sheet))
        currentSheet = theFile[sheet]
        row_list = []
        index = 0
        for row in range(1, currentSheet.max_row + 1):
            for column in "ABCDEF":  # Here you can add or reduce the columns
                cell_name = "{}{}".format(column, row)
                if keyword_text.get().upper() in str(currentSheet[cell_name].value).upper() :
                    row_str = "{} - {} - {} - {} - {} - {}".format(currentSheet["A"+str(row)].value,currentSheet["B"+str(row)].value,\
                                                               currentSheet["C"+str(row)].value,currentSheet["D"+str(row)].value,\
                                                               currentSheet["E"+str(row)].value,currentSheet["F"+str(row)].value)
                    row_list.append(row_str)
                    index += 1
                    results_dic.update({str(sheet):row_list})
                    num_resultados += 1
                    
    # CREATES RESULTS LIST (W/ DUPLICATES)
    results_lista = []
    for key in results_dic:
        for resultado in results_dic[key]:
            results_lista.append(resultado)
    
    # REMOVE DUPLICATES
    results_lista_single = list(collections.OrderedDict.fromkeys(results_lista))
        
    resultados_tklist.delete(0,tk.END)
    for resultados in results_lista_single:
        resultados_tklist.insert(tk.END,resultados)
    
    # DISPLAYS NUMBER OF RESULTS
    num_resultados_msg = tk.Message(app,text="Número de resultados: " + str(len(results_lista_single)))
    num_resultados_msg.grid(row=0,column=3,columnspan=2,sticky="E",pady=15)

    
def abrir_pasta(argv=None):
    num_obra = (resultados_tklist.get(tk.ACTIVE)[:5])
    sec21 = ['20','19','18','17','16','15','14','13','12','11','10','09','07','06','05','04','03','02','01','00']
    if num_obra.startswith(tuple(sec21)):
        nome_ano = 'Ano' + '20' + num_obra[:2]
        path = r"\\fileserver\Obras\Projetos"+ '\\' + str(nome_ano) + '\\' + str(num_obra)
        path = os.path.realpath(path)
        os.startfile(path)
    elif num_obra.startswith(('9')):
        nome_ano = 'Ano' + '19' + num_obra[:2]
        path = r"\\fileserver\Obras\Projetos"+ '\\' + str(nome_ano) + '\\' + str(num_obra)
        path = os.path.realpath(path)
        os.startfile(path)

        
# KEYWORD
keyword_label = tk.Label(app, text='Item a procurar:')
keyword_label.grid(row=0,column=0,sticky="E", pady=15)
keyword_text = tk.StringVar(app)
keyword_entry = tk.Entry(app,textvariable=keyword_text)
keyword_entry.grid(row=0,column=1)


# FIND BUTTON
find_button = tk.Button(app,text='Procurar',width=12,command=find_specific_cell)
find_button.grid(row=0,column=2)
app.bind('<Return>', find_specific_cell)

# ABRIR OBRA BUTTON
abrir_button = tk.Button(app,text='Abrir Pasta',width=12,command=abrir_pasta)
abrir_button.grid(row=0,column=9)


# lISTA DE RESULTADOS
resultados_tklist = tk.Listbox(app,height=20,width=140,border=0)
resultados_tklist.grid(row=3,column=0,columnspan=10,rowspan=10,pady=20,padx=20)
resultados_tklist.bind('<Double-Button-1>',abrir_pasta)
# SCROLLBAR
scrollbar = tk.Scrollbar(app)
scrollbar.grid(row=3,column=10,rowspan=10,sticky="nsw")
# SET SCROLL TO LISTBOX
resultados_tklist.configure(yscrollcommand=scrollbar.set)
scrollbar.configure(command=resultados_tklist.yview)

app.mainloop()
